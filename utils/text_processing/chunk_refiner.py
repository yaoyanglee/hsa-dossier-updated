import os
import logging
import time
import configparser

from alive_progress import alive_bar
from azure.data.tables import TableServiceClient
from azure.core.credentials import AzureNamedKeyCredential
from langchain_community.callbacks import get_openai_callback
from langchain_openai import AzureChatOpenAI
from langchain.prompts import ChatPromptTemplate
from langchain.chains import LLMChain
from tiktoken import encoding_for_model
from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook, Workbook
from utils.table import azure_table_client

# --- Configure logger ---
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
handler = logging.StreamHandler()
formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
handler.setFormatter(formatter)
logger.addHandler(handler)

# # --- Azure table ---

config = configparser.ConfigParser()
config.read("config.prop")
azure_hsa_store_config = config["azure_hsa_store"]
account_name = azure_hsa_store_config["account_name"]
account_key = azure_hsa_store_config["account_key"]

credential = AzureNamedKeyCredential(account_name, account_key)
endpoint = f"https://{account_name}.table.core.windows.net"

table_service_client = TableServiceClient(
    endpoint=endpoint, credential=credential)
table_name = "docmap"

# Load configuration
config = configparser.ConfigParser()
config.read("config.prop")
azure_config = config["azure_openai_gpt4o-mini"]

# Set environment variables
os.environ["AZURE_OPENAI_ENDPOINT"] = azure_config["endpoint"]
os.environ["AZURE_OPENAI_API_KEY"] = azure_config["api_key"]


class ChunkRefiner:
    def __init__(self, deployment_name: str, api_version: str,
                 temperature: float = 0.0, max_retries: int = 3, retry_delay: int = 2,
                 max_tokens_per_section: int = 800, min_tokens_per_section: int = 50):
        
        self.llm = AzureChatOpenAI(
            azure_deployment=deployment_name,
            api_version=api_version,
            temperature=temperature
        )
        self.tokenizer = encoding_for_model("gpt-4o")
        self.max_retries = max_retries
        self.retry_delay = retry_delay
        self.max_tokens_per_section = max_tokens_per_section
        self.min_tokens_per_section = min_tokens_per_section
        self.prompt = ChatPromptTemplate.from_template(
            """You are a precise text refinement and sectioning assistant. Your task is ONLY to:

            Strict Rules:
            1. Break the following text chunk into meaningful sections of maximum {max_tokens} tokens each
            2. Preserve ALL factual information, technical details, citations, and references exactly as they appear.
            3. DO NOT change, modify, or alter any words, characters, punctuations, formatting, or structure—everything must remain exactly the same.
            4. DO NOT paraphrase, rephrase, or summarize.
            5. DO NOT introduce, remove, reorder, or edit any content, including technical terms, numbers, citations, or references.
            6. DO NOT change capitalization, spacing, line breaks, bullet points, or any other formatting.
            7. Preserve all citations, references, and attributions
            8. The structure of each section should follow natural logical breaks (e.g., sentences, paragraphs) without cutting off mid-word, mid-sentence, or mid-citation.
            9. The final output must be word-for-word, character-by-character identical to the input, with the only difference being logical section breaks.
            10. Ensure lossless splitting: Byte-level and character-level consistency must be maintained. Whitespace, tabs, indentation, and newline characters must remain untouched.
            11. The final output must be 100 percent identical to the input at a token, word, character, and formatting level, except for the section breaks.
            
            Format each section as:

            [SECTION_START]
            <section content>
            [SECTION_END]
            
            Text chunk to refine and section:
            {chunk}
            
            Refined and sectioned text:"""
        )
        self.chain = LLMChain(llm=self.llm, prompt=self.prompt)


    def get_document_metadata(self, filename: str, project_name: str, doc_name: str, subfolder_type: str) -> str:
        """
        Fetches the hashed document name from Azure Table and determines the subfolder type.
        
        Args:
            filename (str): The document filename.
            project_name (str): The project name (e.g., folder name under docs).
            source_folder (str): Indicates the source folder type (Clinical literature or User manual).
        
        Returns:
            str: The refined filename suffix (hashed_doc_name-lr or hashed_doc_name-ifu).
        """
        
        entity = azure_table_client.retrieve_by_doc_name(
            table_name="docmap",
            project_name=project_name,
            doc_name=doc_name
        )
        if entity:
            # Entity is found, assign hashed_doc
            hashed_doc = entity
        else:
            # Handle the case where entity is not found, e.g., raise an exception or log the error
            raise ValueError(f"No matching entry found in Azure Table for {filename}")

        # Combine the hashed document with subfolder type
        return f"{hashed_doc}-{subfolder_type}"
    

    def count_tokens(self, text: str) -> int:
        """Count tokens in text using tiktoken"""
        return len(self.tokenizer.encode(text))
    

    def parse_sections(self, refined_text: str) -> List[str]:
        """Parse sections from refined text"""
        sections = []
        current_section = []
        
        for line in refined_text.split('\n'):
            if line.strip() == '[SECTION_START]':
                current_section = []
            elif line.strip() == '[SECTION_END]':
                if current_section:
                    sections.append('\n'.join(current_section).strip())
            else:
                current_section.append(line)
        
        return sections
    

    def refine_chunk(self, project_name: str, chunk: str, chunk_index: int, refined_filename_suffix: str,filename:str) -> List[Dict]:
        """
        Refine a single chunk into sections with retry logic.
        
        Args:
            chunk: Text chunk to refine.
            chunk_index: Index of the chunk.
            refined_filename_suffix: The suffix to append to each section filename.
            
        Returns:
            List of dictionaries containing refined sections and their metadata.
        """
        print("filename", filename)

        refined_sections = []
        for attempt in range(self.max_retries):
            try:
                refined = self.chain.run(chunk=chunk, max_tokens=self.max_tokens_per_section)

                
                sections = self.parse_sections(refined.strip())
                #added here
                combined__refined_text = "\n\n".join(sections)
                self.save_to_excel(project_name, filename, chunk,combined__refined_text,chunk_index)
                for section_index, section in enumerate(sections, 1):
                    section_id = f"{project_name}-{refined_filename_suffix}-chunk{chunk_index}-section{section_index}"
                    token_count = self.count_tokens(section)
                    if token_count > self.min_tokens_per_section:
                        refined_sections.append({
                            'section_id': section_id,
                            'content': section,
                            'token_count': token_count
                        })
                return refined_sections  # Return once successful
            except Exception as e:
                logger.error(f"Error refining chunk {chunk_index} on attempt {attempt + 1}/{self.max_retries}: {e}")
                if attempt == self.max_retries - 1:
                    logger.warning(f"Skipping chunk {chunk_index} after {self.max_retries} failed attempts.")
        return refined_sections  # Return whatever sections were successfully refined (or empty if none)
    

    def refine_chunks_and_save(self, chunks: List, filename: str, project_name: str, document_name: str, subfolder_type: str, output_dir="text_sections"):
        """
        Refine chunks into sections and save them with the new naming convention.
        
        Args:
            chunks: List of text chunks to refine.
            filename: Original filename.
            project_name: Name of the project (partition key for Azure Table).
            source_folder: Source folder name (determines if it's 'lr' or 'ifu').
            output_dir: Directory to save the refined sections.
        """
        # Get document metadata    
        logger.info(f"*** Refining chunks for: {document_name}")
        refined_filename_suffix = self.get_document_metadata(filename, project_name, document_name, subfolder_type)
        os.makedirs(output_dir, exist_ok=True)

        logger.info(f"Starting refinement of {len(chunks)} chunks for file {filename}...")

        with get_openai_callback() as cb:
            with alive_bar(len(chunks), title="Refining chunks") as bar:
                for chunk_index, chunk in enumerate(chunks, 1):
                    sections = self.refine_chunk(project_name, chunk, chunk_index, refined_filename_suffix,filename)
                    for section in sections:
                        section_filename = f"{section['section_id']}.txt"
                        section_path = os.path.join(output_dir, section_filename)
                        with open(section_path, "w", encoding="utf-8") as f:
                            f.write(section['content'])
                    bar()

        logger.info(f"Refinement complete. All sections saved in {output_dir}")




    def save_to_excel(self,project_name: str, filename: str, original_chunk: str, refined_chunk: str,chunk_index:int):
        """
        Save original and refined chunks into an Excel file, appending to the next available row.

        Args:
            project_name: Name of the project (used for the Excel filename).
            filename: Name of the file (used for the sheet name).
            original_chunk: The original chunk before refinement.
            refined_chunk: The refined chunk.
        """
        excel_filename = f"{project_name}_refined_chunks_comparison.xlsx"  # Different file for each project
            # Normalize path to handle backslashes and forward slashes
        normalized_filename = os.path.normpath(filename)

        # Extract the last part (file name without extension)
        last_part = os.path.splitext(os.path.basename(normalized_filename))[0]
        sheet_name = last_part[:31]  # Excel sheet names have a 31-character limit
        print("Filename:", filename)
        print("Sheet name:", sheet_name)   

        # Append data properly to the next available row
        self.append_to_next_row(excel_filename, sheet_name,str(original_chunk), refined_chunk)
        print(f"Saved refined chunk to {excel_filename}, sheet: {sheet_name}")




    def append_to_next_row(self, excel_filename, sheet_name, original_chunk, refined_chunk):
        try:
            # Load the existing workbook
            if os.path.exists(excel_filename):
                wb = load_workbook(excel_filename)
            else:
                print(f"File '{excel_filename}' not found. Creating a new file.")
                wb = Workbook()
            
            # Check if sheet exists, create if not
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(["Original Chunk", "Refined Chunk"])  # Add headers
            
            ws = wb[sheet_name]

            # Find the next empty row
            next_row = ws.max_row + 1
            print(f"Appending data to row {next_row}")

            # Append data directly (without DataFrame)
            ws.append([original_chunk, refined_chunk])

            # Save the workbook
            wb.save(excel_filename)
            print("Data successfully appended to the next row.")

        except Exception as e:
            print(f"Error: {e}")